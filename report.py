#!/usr/bin/python3

import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, PatternFill

wb = load_workbook('Virtual_QA.xlsx')
ws1 = wb["Report"]
ws2 = wb["QA_checks"]
pf = PatternFill(fgColor="F70825", patternType = "solid")
font_p = Font(bold=True, color="00FFFFFF")
pp = PatternFill(fgColor="6DF111", patternType = "solid")
#OS info
ws1['C3'] = (sys.argv[1])
ws1['C4'] = (sys.argv[2])
ws1['C5'] = (sys.argv[3])
ws1['C6'] = (sys.argv[4])

#Processor info
ws1['C9'] = (sys.argv[5])

#Memory info
ws1['C12'] = (sys.argv[6])

#Disk info
ws1['C15'] = (sys.argv[7])

#Network info
ws1['C18'] = (sys.argv[8])
ws1['C19'] = (sys.argv[9])
ws1['C20'] = (sys.argv[10])

#Red Hat Subscription info
if "not" in (sys.argv[30]):
    ws1['C24'] = (sys.argv[30])
    ws2['D19'].fill = pf
    ws2['D19'].font = font_pf
    ws2['D19'] = "FAIL"
    ws2['E19'] = (sys.argv[30])
elif "false" in (sys.argv[31]):
    ws1['C24'] = (sys.argv[30])
    ws2['D19'].fill = pf
    ws2['D19'].font = font_pf
    ws2['D19'] = "FAIL"
else:
    ws1['C24'] = (sys.argv[30])
    ws2['D19'].fill = pp
    ws2['D19'].font = font_pf
    ws2['D19'] = "PASS"

#AD info
if "not" in (sys.argv[11]):
    ws1['C27'].fill = pf
    ws1['C27'].font = font_p
    ws1['C27'] = "Centrify Agent Not Installed"

#else:
    ws1['C27'] = (sys.argv[29])

#Services info


if "yes" in (sys.argv[12]):
    ws1['C30'] = "Active"
    ws2['D18'].fill = pp
    ws2['D18'].font = font_p
    ws2['D18'] = "PASS"
else:
    ws1['C30'].fill = pf
    ws1['C30'].font = font_p
    ws1['C30'] = "Inactive"
    ws2['D18'].fill = pf
    ws2['D18'].font = font_pf
    ws2['D18'] = "FAIL"
    ws2['E18'] = "NTP not Running"
    
if "running" in (sys.argv[14]):
    ws1['C32'] = "Active"
    ws2['D21'].fill = pp
    ws2['D21'].font = font_p
    ws2['D21'] = "PASS"
else:
    ws1['C32'].fill = pf
    ws1['C32'].font = font_p
    ws1['C32'] = "Inactive"
    ws2['D21'].fill = pf
    ws2['D21'].font = font_pf
    ws2['D21'] = "FAIL"
    ws2['E21'] = "Monitoring Agent " + (sys.argv[13]) +  " not Running"
    
if "running" in (sys.argv[15]):
    ws1['C33'] = "Active"
    ws2['D22'].fill = pp
    ws2['D22'].font = font_p
    ws2['D22'] = "PASS"
else:
    ws1['C33'].fill = pf
    ws1['C33'].font = font_p
    ws1['C33'] = "Inactive"
    ws2['D22'].fill = pp
    ws2['D22'].font = font_pf
    ws2['D22'] = "FAIL"
    ws2['E22'] = "Backup agent not Running"
    
if "running" in (sys.argv[16]):
    ws1['C35'] = "Active"
    ws2['D23'].fill = pp
    ws2['D23'].font = font_p
    ws2['D23'] = "PASS"
else:
    ws1['C35'].fill = pf
    ws1['C35'].font = font_p
    ws1['C35'] = "Inactive"
    ws2['D23'].fill = pf
    ws2['D23'].font = font_pf
    ws2['D23'] = "FAIL"
    ws2['E23'] = "DS agent not Running"
    
if "running" in (sys.argv[17]):
    ws1['C34'] = "Active"
    ws2['D24'].fill = pp
    ws2['D24'].font = font_p
    ws2['D24'] = "PASS"    
else:
    ws1['C34'].fill = pf
    ws1['C34'].font = font_p
    ws1['C34'] = "Inactive"
    ws2['D24'].fill = pf
    ws2['D24'].font = font_pf
    ws2['D24'] = "FAIL"
    ws2['E24'] = "VMware Tools not installed"
    
if "running" in (sys.argv[18]):
    ws1['C31'] = "Active"
    ws2['D20'].fill = pp
    ws2['D20'].font = font_p
    ws2['D20'] = "PASS"
else:
    ws1['C31'].fill = pf
    ws1['C31'].font = font_p
    ws1['C31'] = "Inactive"    
    ws2['D20'].fill = pf
    ws2['D20'].font = font_pf
    ws2['D20'] = "FAIL"
    ws2['E20'] = "KDump Agent not running"
    
    
ws2['C32'] = (sys.argv[19])
ws2['C33'] = (sys.argv[20])
ws2['C34'] = (sys.argv[21])
ws2['C35'] = (sys.argv[22])

if "true" in (sys.argv[23]):
    ws2['D6'].fill = pp
    ws2['D6'].font = font_p
    ws2['D6'] = "PASS"
    ws2['D7'].fill = pp
    ws2['D7'].font = font_p
    ws2['D7'] = "PASS"
    ws2['D8'].fill = pp
    ws2['D8'].font = font_p
    ws2['D8'] = "PASS"
else:
    ws2['D6'].fill = pf
    ws2['D6'].font = font_p
    ws2['D6'] = "FAIL"
    ws2['D7'].fill = pf
    ws2['D7'].font = font_p
    ws2['D7'] = "FAIL"
    ws2['D8'].fill = pf
    ws2['D8'].font = font_p
    ws2['D8'] = "FAIL"
    
if "true" in (sys.argv[24]):
    ws2['D9'].fill = pp
    ws2['D9'].font = font_p
    ws2['D9'] = "PASS"
    ws2['D10'].fill = pp
    ws2['D10'].font = font_p
    ws2['D10'] = "PASS"
else:
    ws2['D9'].fill = pf
    ws2['D9'].font = font_p
    ws2['D9'] = "FAIL"
    ws2['D10'].fill = pf
    ws2['D10'].font = font_p
    ws2['D10'] = "FAIL"

if "true" in (sys.argv[25]):
    ws2['D11'].fill = pp
    ws2['D11'].font = font_p
    ws2['D11'] = "PASS"
else:
    ws2['D11'].fill = pf
    ws2['D11'].font = font_p
    ws2['D11'] = "FAIL"
    
if "true" in (sys.argv[26]):
    ws2['D12'].fill = pp
    ws2['D12'].font = font_p
    ws2['D12'] = "PASS"
    ws2['D13'].fill = pp
    ws2['D13'].font = font_p
    ws2['D13'] = "PASS"
    ws2['D14'].fill = pp
    ws2['D14'].font = font_p
    ws2['D14'] = "PASS"
else:
    ws2['D12'].fill = pf
    ws2['D12'].font = font_p
    ws2['D12'] = "FAIL"
    ws2['D13'].fill = pf
    ws2['D13'].font = font_p
    ws2['D13'] = "FAIL"
    ws2['D14'].fill = pf
    ws2['D14'].font = font_p
    ws2['D14'] = "FAIL"
    
if "failed" in (sys.argv[11]):
    ws2['E15'] = "Centrify Agent Not Installed/Running"
    ws2['D15'].fill = pf
    ws2['D15'].font = font_p
    ws2['D15'] = "FAIL"

else:
    ws2['D15'].fill = pp
    ws2['D15'].font = font_p
    ws2['D15'] = "PASS"

    ws2['E16'] = (sys.argv[28])
if "true" in (sys.argv[27]):
    ws2['D16'].fill = pp
    ws2['D16'].font = font_p
    ws2['D16'] = "PASS"
    ws2['D17'].fill = pp
    ws2['D17'].font = font_p
    ws2['D17'] = "PASS"
else:
    ws2['D16'].fill = pf
    ws2['D16'].font = font_p
    ws2['D16'] = "FAIL"
    ws2['D17'].fill = pf
    ws2['D17'].font = font_p
    ws2['D17'] = "FAIL"

     
wb.save('test.xlsx')
