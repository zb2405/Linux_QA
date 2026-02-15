#!/usr/bin/python3

import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# -----------------------------
# Helpers
# -----------------------------

def arg(i, default=""):
    try:
        return sys.argv[i]
    except IndexError:
        return default

def is_true(val):
    return str(val).lower() in ["true", "yes", "1"]

def contains(text, key):
    return key.lower() in str(text).lower()

# -----------------------------
# Styles
# -----------------------------

FAIL_FILL = PatternFill(fgColor="F70825", patternType="solid")
PASS_FILL = PatternFill(fgColor="6DF111", patternType="solid")

FONT_WHITE = Font(bold=True, color="00FFFFFF")
FONT_BOLD = Font(bold=True)

# -----------------------------
# Load workbook
# -----------------------------

wb = load_workbook("Virtual_QA.xlsx")
ws1 = wb["Report"]
ws2 = wb["QA_checks"]

# -----------------------------
# Arguments Mapping
# -----------------------------

hostname = arg(1)
os_name = arg(2)
os_version = arg(3)
srv_type = arg(4)
vcpu = arg(5)
ram = arg(6)
disk = arg(7)
ip = arg(8)
mask = arg(9)
gw = arg(10)
yum_cmd = arg(11)
ntp = arg(12)
mon_agent = arg(13)
mon_status = arg(14)
backup_status = arg(15)
ds_status = arg(16)
vmtools_status = arg(17)
kdump_status = arg(18)
build_by = arg(19)
qa_by = arg(20)
ticket = arg(21)
client = arg(22)

os_ok = arg(23)
proc_ok = arg(24)
sto_ok = arg(25)
nw_ok = arg(26)
dns_ok = arg(27)

dns_domain = arg(28)
ad_info = arg(29)
osr_out = arg(30)
osr_verify = arg(31)

# -----------------------------
# Report Sheet
# -----------------------------

ws1["C3"] = hostname
ws1["C4"] = os_name
ws1["C5"] = os_version
ws1["C6"] = srv_type

ws1["C9"] = vcpu
ws1["C12"] = ram
ws1["C15"] = disk

ws1["C18"] = ip
ws1["C19"] = mask
ws1["C20"] = gw

# -----------------------------
# OS Registration Logic
# -----------------------------

ws1["C24"] = osr_out

if contains(osr_out, "not") or contains(osr_verify, "false"):
    ws2["D19"].fill = FAIL_FILL
    ws2["D19"].font = FONT_WHITE
    ws2["D19"] = "FAIL"
else:
    ws2["D19"].fill = PASS_FILL
    ws2["D19"].font = FONT_WHITE
    ws2["D19"] = "PASS"

# -----------------------------
# AD Info
# -----------------------------

ws1["C27"] = ad_info

# -----------------------------
# Services Checks
# -----------------------------

def service_check(cell_report, cell_qa, status, fail_msg):
    if contains(status, "running") or contains(status, "active"):
        ws1[cell_report] = "Active"
        ws2[cell_qa].fill = PASS_FILL
        ws2[cell_qa] = "PASS"
    else:
        ws1[cell_report].fill = FAIL_FILL
        ws1[cell_report] = "Inactive"
        ws2[cell_qa].fill = FAIL_FILL
        ws2[cell_qa] = "FAIL"

service_check("C30", "D18", ntp, "NTP not running")
service_check("C32", "D21", mon_status, f"{mon_agent} not running")
service_check("C33", "D22", backup_status, "Backup not running")
service_check("C35", "D23", ds_status, "DS agent not running")
service_check("C34", "D24", vmtools_status, "VM tools not running")
service_check("C31", "D20", kdump_status, "Kdump not running")

# -----------------------------
# QA Pass/Fail Mapping
# -----------------------------

def qa_block(start_row, ok):
    fill = PASS_FILL if is_true(ok) else FAIL_FILL
    text = "PASS" if is_true(ok) else "FAIL"
    for r in start_row:
        ws2[r].fill = fill
        ws2[r] = text

qa_block(["D6","D7","D8"], os_ok)
qa_block(["D9","D10"], proc_ok)
qa_block(["D11"], sto_ok)
qa_block(["D12","D13","D14"], nw_ok)
qa_block(["D16","D17"], dns_ok)

# -----------------------------
# Footer Info
# -----------------------------

ws2["C32"] = build_by
ws2["C33"] = qa_by
ws2["C34"] = ticket
ws2["C35"] = client

wb.save("test.xlsx")
